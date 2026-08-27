# -*- coding: utf-8 -*-
"""Отчёт по всем ботам в Excel: по дням — сделки, winrate, PnL.

Запуск: python -m backend.report_excel [путь.xlsx]
Выход по умолчанию: data/bots_report.xlsx
"""
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import storage

BOTS = [
    ("Гибрид (лимитки+SL)", "hybrid"),
    ("Скальпер V2 (EMA+ADX)", ""),
    ("Мульти-бот (6 инстр.)", "multi"),
    ("Скальпер V3 (ML+фигуры)", "kiro"),
    ("Сетка (grid)", "grid"),
]
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
WIN_FILL = PatternFill("solid", fgColor="DCFCE7")
LOSS_FILL = PatternFill("solid", fgColor="FEE2E2")
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="374151")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def day_stats(version: str) -> dict:
    """{дата: {trades, wins, pnl}} по закрытым сделкам версии."""
    if version == "grid":
        q = ("SELECT substr(close_time,1,10) d, pnl FROM pairs "
             "WHERE status='closed'")
        rows = _fetch(q)
    else:
        cond = "(version='' OR version IS NULL)" if version == "" \
            else f"version='{version}'"
        q = (f"SELECT substr(close_time,1,10) d, pnl FROM scalp_trades "
             f"WHERE status='closed' AND {cond}")
        rows = _fetch(q)
    out = defaultdict(lambda: {"trades": 0, "wins": 0, "pnl": 0.0})
    for d, pnl in rows:
        out[d]["trades"] += 1
        out[d]["pnl"] += pnl or 0.0
        if pnl and pnl > 0:
            out[d]["wins"] += 1
    return dict(out)


def _fetch(q):
    with storage.get_conn() as c:
        return [(r[0], r[1]) for r in c.execute(q)]


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def build(path: Path):
    storage.init_db()
    wb = Workbook()

    # ---------- Лист 1: сводка по ботам ----------
    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Бот", "Сделок всего", "Winrate %", "PnL всего $", "Дней в работе",
               "Сделок/день", "PnL/день $"])
    style_header(ws, 1, 7)
    row = 2
    for name, ver in BOTS:
        ds = day_stats(ver)
        trades = sum(d["trades"] for d in ds.values())
        wins = sum(d["wins"] for d in ds.values())
        pnl = sum(d["pnl"] for d in ds.values())
        days = len(ds)
        ws.append([name, trades,
                   round(100 * wins / trades, 1) if trades else 0,
                   round(pnl, 2),
                   days,
                   round(trades / days, 1) if days else 0,
                   round(pnl / days, 2) if days else 0])
        pnl_cell = ws.cell(row=row, column=4)
        pnl_cell.fill = WIN_FILL if pnl >= 0 else LOSS_FILL
        row += 1
    for col, w in zip("ABCDEFG", [26, 13, 11, 12, 13, 12, 12]):
        ws.column_dimensions[col].width = w

    # ---------- Лист 2: по дням ----------
    ws2 = wb.create_sheet("По дням")
    all_days = sorted({d for _, ver in BOTS for d in day_stats(ver)})
    ws2.append(["Дата"] + [n for n, _ in BOTS] + ["Итого за день $"])
    style_header(ws2, 1, 2 + len(BOTS))
    grand = defaultdict(float)
    for r, day in enumerate(all_days, start=2):
        ws2.cell(row=r, column=1, value=day)
        total_day = 0.0
        for ci, (name, ver) in enumerate(BOTS, start=2):
            d = day_stats(ver).get(day)
            if d:
                ws2.cell(row=r, column=ci,
                         value=f"{d['trades']} сд / WR "
                               f"{round(100 * d['wins'] / d['trades'])}% / "
                               f"{round(d['pnl'], 2):+}$")
                total_day += d["pnl"]
                grand[name] += d["pnl"]
            else:
                ws2.cell(row=r, column=ci, value="—")
        c = ws2.cell(row=r, column=2 + len(BOTS), value=round(total_day, 2))
        c.font = TOTAL_FONT
        c.fill = WIN_FILL if total_day >= 0 else LOSS_FILL
    # итоговая строка
    r = len(all_days) + 2
    ws2.cell(row=r, column=1, value="ИТОГО").font = TOTAL_FONT
    for ci, (name, _) in enumerate(BOTS, start=2):
        c = ws2.cell(row=r, column=ci, value=round(grand[name], 2))
        c.font = TOTAL_FONT
        c.fill = WIN_FILL if grand[name] >= 0 else LOSS_FILL
    for col in range(1, 2 + len(BOTS) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22
        for rr in range(1, r + 1):
            ws2.cell(row=rr, column=col).border = BORDER

    # ---------- Лист 3: все сделки ----------
    ws3 = wb.create_sheet("Все сделки")
    ws3.append(["Бот", "Символ", "Сторона", "Вход", "Выход", "Открыта",
                "Закрыта", "Причина", "PnL $"])
    style_header(ws3, 1, 9)
    r = 2
    for name, ver in BOTS:
        if ver == "grid":
            q = ("SELECT 'XAUUSD.s', 'LONG', buy_price, sell_price, open_time, "
                 "close_time, 'tp', pnl FROM pairs WHERE status='closed'")
        else:
            cond = "(version='' OR version IS NULL)" if ver == "" \
                else f"version='{ver}'"
            q = (f"SELECT symbol, side, entry, exit, open_time, close_time, "
                 f"reason, pnl FROM scalp_trades WHERE status='closed' AND {cond}")
        with storage.get_conn() as c:
            for row_vals in c.execute(q):
                ws3.append([name, *row_vals])
                pnl_cell = ws3.cell(row=r, column=9)
                pnl_cell.fill = (WIN_FILL if (row_vals[7] or 0) >= 0
                                 else LOSS_FILL)
                r += 1
    for col, w in zip("ABCDEFGHI", [22, 11, 9, 10, 10, 19, 19, 8, 9]):
        ws3.column_dimensions[col].width = w

    wb.save(path)
    print(f"готово: {path}")
    print(f"листы: Сводка · По дням ({len(all_days)} дней) · Все сделки")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent.parent / "data" / "bots_report.xlsx"
    build(out)
